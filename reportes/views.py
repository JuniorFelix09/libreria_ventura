import io
from datetime import datetime, timedelta

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.utils import timezone

from cuentas.decorators import solo_admin
from caja.models import Turno, MovimientoCaja
from ventas.models import Venta
from productos.models import Producto

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


@login_required
@solo_admin
def index(request):
    turnos_cerrados = Turno.objects.filter(
        estado=Turno.Estado.CERRADO
    ).select_related('caja', 'cajero').order_by('-fecha_cierre')[:30]

    hoy = timezone.localdate()
    hace_30 = hoy - timedelta(days=30)

    return render(request, 'reportes/index.html', {
        'turnos_cerrados': turnos_cerrados,
        'fecha_inicio_default': hace_30.isoformat(),
        'fecha_fin_default': hoy.isoformat(),
    })


# ---------------------------------------------------------------------------
# PDF: cierre de turno
# ---------------------------------------------------------------------------

@login_required
@solo_admin
def cierre_turno_pdf(request, pk):
    turno = get_object_or_404(Turno, pk=pk)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        topMargin=2 * cm, bottomMargin=2 * cm,
        leftMargin=2 * cm, rightMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle(
        'TituloReporte', parent=styles['Title'], fontSize=16,
        textColor=colors.HexColor('#1a237e'), spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        'SubReporte', parent=styles['Normal'], fontSize=9,
        textColor=colors.HexColor('#666666'), spaceAfter=16,
    )
    seccion_style = ParagraphStyle(
        'Seccion', parent=styles['Heading2'], fontSize=12,
        textColor=colors.HexColor('#1a237e'), spaceBefore=14, spaceAfter=6,
    )

    story = []
    story.append(Paragraph('Librería Ventura — Cierre de turno', titulo_style))
    story.append(Paragraph(
        f'{turno.caja} · Cajero: {turno.cajero.get_full_name() or turno.cajero.username} · '
        f'Generado el {timezone.localtime():%d/%m/%Y %H:%M}',
        sub_style,
    ))

    estado_txt = turno.get_estado_display()
    cierre_txt = (
        timezone.localtime(turno.fecha_cierre).strftime('%d/%m/%Y %H:%M')
        if turno.fecha_cierre else '—'
    )
    datos = [
        ['Apertura', timezone.localtime(turno.fecha_apertura).strftime('%d/%m/%Y %H:%M')],
        ['Cierre', cierre_txt],
        ['Estado', estado_txt],
        ['Monto inicial', f'RD${turno.monto_inicial}'],
    ]
    t_datos = Table(datos, colWidths=[6 * cm, 6 * cm])
    t_datos.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#666666')),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#e8eaf0')),
    ]))
    story.append(t_datos)

    story.append(Paragraph('Ventas del turno', seccion_style))
    ventas_qs = turno.ventas.filter(anulada=False).order_by('fecha')
    total_efectivo = sum(
        v.total for v in ventas_qs if v.metodo_pago == Venta.MetodoPago.EFECTIVO
    )
    total_tarjeta = sum(
        v.total for v in ventas_qs if v.metodo_pago == Venta.MetodoPago.TARJETA
    )

    resumen_ventas = [
        ['Método de pago', 'Cantidad', 'Total'],
        ['Efectivo', str(ventas_qs.filter(metodo_pago=Venta.MetodoPago.EFECTIVO).count()), f'RD${total_efectivo}'],
        ['Tarjeta', str(ventas_qs.filter(metodo_pago=Venta.MetodoPago.TARJETA).count()), f'RD${total_tarjeta}'],
    ]
    t_resumen = Table(resumen_ventas, colWidths=[6 * cm, 3 * cm, 5 * cm])
    t_resumen.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e8eaf0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7ff')]),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(t_resumen)

    movimientos = turno.movimientos.all().order_by('fecha')
    if movimientos:
        story.append(Paragraph('Movimientos de caja', seccion_style))
        filas_mov = [['Fecha', 'Tipo', 'Motivo', 'Monto']]
        for m in movimientos:
            signo = '+' if m.tipo == MovimientoCaja.Tipo.INGRESO else '-'
            filas_mov.append([
                timezone.localtime(m.fecha).strftime('%d/%m %H:%M'),
                m.get_tipo_display(),
                m.motivo,
                f'{signo}RD${m.monto}',
            ])
        t_mov = Table(filas_mov, colWidths=[3 * cm, 3.5 * cm, 5.5 * cm, 2.5 * cm])
        t_mov.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e8eaf0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7ff')]),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(t_mov)

    if turno.estado == Turno.Estado.CERRADO:
        story.append(Paragraph('Cierre de caja', seccion_style))
        diferencia = turno.diferencia or 0
        if diferencia > 0:
            estado_dif, color_dif = 'Sobrante', colors.HexColor('#2e7d32')
        elif diferencia < 0:
            estado_dif, color_dif = 'Faltante', colors.HexColor('#c62828')
        else:
            estado_dif, color_dif = 'Exacto', colors.HexColor('#2e7d32')

        filas_cierre = [
            ['Esperado en efectivo', f'RD${turno.monto_esperado_efectivo}'],
            ['Contado físico', f'RD${turno.monto_contado}'],
            ['Diferencia', f'RD${diferencia} ({estado_dif})'],
        ]
        t_cierre = Table(filas_cierre, colWidths=[6 * cm, 6 * cm])
        t_cierre.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#666666')),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (1, 2), (1, 2), color_dif),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#e8eaf0')),
        ]))
        story.append(t_cierre)
    else:
        story.append(Spacer(1, 10))
        story.append(Paragraph('Este turno todavía está abierto.', styles['Italic']))

    doc.build(story)
    buffer.seek(0)

    filename = f'cierre_turno_{turno.pk}_{timezone.localdate():%Y%m%d}.pdf'
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# Excel: ventas por rango de fechas
# ---------------------------------------------------------------------------

@login_required
@solo_admin
def ventas_excel(request):
    hoy = timezone.localdate()
    fecha_inicio = request.GET.get('fecha_inicio') or (hoy - timedelta(days=30)).isoformat()
    fecha_fin = request.GET.get('fecha_fin') or hoy.isoformat()

    inicio_dt = timezone.make_aware(datetime.strptime(fecha_inicio, '%Y-%m-%d'))
    fin_dt = timezone.make_aware(
        datetime.strptime(fecha_fin, '%Y-%m-%d') + timedelta(days=1)
    )

    ventas = (
        Venta.objects
        .filter(fecha__gte=inicio_dt, fecha__lt=fin_dt)
        .select_related('cajero', 'turno', 'turno__caja')
        .prefetch_related('detalles', 'detalles__producto')
        .order_by('fecha')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Ventas'

    encabezado_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    encabezado_font = Font(color='FFFFFF', bold=True)

    headers = ['Venta #', 'Fecha', 'Caja', 'Cajero', 'Producto', 'Cantidad', 'Precio unit.', 'Subtotal', 'Método de pago', 'Estado']
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        celda = ws.cell(row=1, column=col)
        celda.fill = encabezado_fill
        celda.font = encabezado_font
        celda.alignment = Alignment(horizontal='center')

    total_general = 0
    for venta in ventas:
        estado = 'Anulada' if venta.anulada else 'Válida'
        detalles = venta.detalles.all()
        if not detalles:
            continue
        for det in detalles:
            ws.append([
                venta.pk,
                timezone.localtime(venta.fecha).strftime('%d/%m/%Y %H:%M'),
                venta.turno.caja.nombre,
                venta.cajero.get_full_name() or venta.cajero.username,
                det.producto.nombre,
                det.cantidad,
                float(det.precio_unitario),
                float(det.subtotal),
                venta.get_metodo_pago_display(),
                estado,
            ])
        if not venta.anulada:
            total_general += venta.total

    ws.append([])
    ws.append(['', '', '', '', '', '', '', 'Total período', float(total_general), ''])
    fila_total = ws.max_row
    ws.cell(row=fila_total, column=8).font = Font(bold=True)
    ws.cell(row=fila_total, column=9).font = Font(bold=True)

    anchos = [10, 17, 12, 18, 26, 10, 12, 12, 15, 10]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.freeze_panes = 'A2'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'ventas_{fecha_inicio}_a_{fecha_fin}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# Excel: inventario / stock
# ---------------------------------------------------------------------------

@login_required
@solo_admin
def inventario_excel(request):
    productos = Producto.objects.select_related('categoria').order_by('categoria__nombre', 'nombre')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'

    encabezado_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    encabezado_font = Font(color='FFFFFF', bold=True)
    alerta_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')

    headers = ['Producto', 'Categoría', 'Precio', 'Stock', 'Valor en inventario', 'Activo']
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        celda = ws.cell(row=1, column=col)
        celda.fill = encabezado_fill
        celda.font = encabezado_font
        celda.alignment = Alignment(horizontal='center')

    STOCK_BAJO = 5
    valor_total = 0
    for p in productos:
        valor_linea = float(p.precio) * p.stock
        valor_total += valor_linea
        ws.append([
            p.nombre,
            p.categoria.nombre if p.categoria else '—',
            float(p.precio),
            p.stock,
            valor_linea,
            'Sí' if p.activo else 'No',
        ])
        if p.stock <= STOCK_BAJO:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = alerta_fill

    ws.append([])
    ws.append(['', '', '', 'Valor total', valor_total, ''])
    fila_total = ws.max_row
    ws.cell(row=fila_total, column=4).font = Font(bold=True)
    ws.cell(row=fila_total, column=5).font = Font(bold=True)

    anchos = [28, 18, 12, 10, 18, 10]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.freeze_panes = 'A2'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'inventario_{timezone.localdate():%Y%m%d}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response